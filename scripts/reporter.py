import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import yaml
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))


def run_reporter(kpi_results):

    # Load config
    with open("config.yaml", "r") as f:
        config = yaml.safe_load(f)

    output_folder = config["reporting"]["output_folder"]
    os.makedirs(output_folder, exist_ok=True)

    report_date = datetime.now().strftime("%Y-%m-%d")
    filename = f"{output_folder}OpsPulse_Report_{report_date}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI Report"

    ws.sheet_view.showGridLines = False

    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    alert_fill = PatternFill("solid", fgColor="FF0000")
    warning_fill = PatternFill("solid", fgColor="FFC000")
    good_fill = PatternFill("solid", fgColor="70AD47")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(cell, text):
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    def style_cell(cell, value, fill=None, align=center):
        cell.value = value
        cell.alignment = align
        cell.border = thin
        if fill:
            cell.fill = fill

    # --- Title ---
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"OpsPulse KPI Report — Week of "
        f"{kpi_results['current_week_start']} to {kpi_results['max_date']}"
    )
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    # --- KPI Header ---
    headers = ["KPI", "Current Week", "Previous Week", "Growth %", "Status"]
    for col, header in zip(["A", "B", "C", "D", "E"], headers):
        style_header(ws[f"{col}3"], header)

    # --- KPI Rows ---
    kpi_rows = [
        ("Total Revenue", kpi_results["current_kpis"]["total_revenue"],
         kpi_results["previous_kpis"]["total_revenue"], kpi_results["growth"]["revenue_growth"]),
        ("Units Sold", kpi_results["current_kpis"]["units_sold"],
         kpi_results["previous_kpis"]["units_sold"], kpi_results["growth"]["units_growth"]),
        ("Avg Order Value", kpi_results["current_kpis"]["aov"],
         kpi_results["previous_kpis"]["aov"], kpi_results["growth"]["aov_growth"]),
        ("Customer Count", kpi_results["current_kpis"]["customer_count"],
         kpi_results["previous_kpis"]["customer_count"], None),
        ("Transaction Count", kpi_results["current_kpis"]["transaction_count"],
         kpi_results["previous_kpis"]["transaction_count"], None),
    ]

    for i, (label, curr, prev, growth) in enumerate(kpi_rows, start=4):

        style_cell(ws[f"A{i}"], label, align=left)
        style_cell(ws[f"B{i}"], curr, align=right)
        style_cell(ws[f"C{i}"], prev, align=right)

        # Currency formatting
        if label in ["Total Revenue", "Avg Order Value"]:
            ws[f"B{i}"].number_format = '"$"#,##0.00'
            ws[f"C{i}"].number_format = '"$"#,##0.00'

        # Growth handling
        if growth is not None:
            ws[f"D{i}"] = growth / 100
            ws[f"D{i}"].number_format = "0.0%"
            ws[f"D{i}"].alignment = right
            ws[f"D{i}"].border = thin

            if growth <= -10:
                fill = alert_fill
                status = "ALERT"
            elif growth <= -5:
                fill = warning_fill
                status = "WARNING"
            else:
                fill = good_fill
                status = "OK"

            ws[f"D{i}"].fill = fill
            style_cell(ws[f"E{i}"], status, fill, align=center)
        else:
            ws[f"D{i}"].border = thin
            ws[f"E{i}"].border = thin

    # --- Alerts Section ---
    alert_row = 11
    ws.merge_cells(f"A{alert_row}:E{alert_row}")
    style_header(ws[f"A{alert_row}"], "ALERTS")

    for j, alert in enumerate(kpi_results["alerts"], start=alert_row + 1):
        ws.merge_cells(f"A{j}:E{j}")
        cell = ws[f"A{j}"]
        cell.value = alert
        cell.alignment = left
        cell.border = thin

        if "ALERT" in alert:
            cell.fill = alert_fill
            cell.font = Font(color="FFFFFF", bold=True)
        else:
            cell.fill = good_fill

    # --- Top Categories Section ---
    top_row = alert_row + len(kpi_results["alerts"]) + 3
    ws.merge_cells(f"A{top_row}:B{top_row}")
    style_header(ws[f"A{top_row}"], "TOP CATEGORIES BY REVENUE")

    for k, row in kpi_results["top_products"].iterrows():
        r = top_row + k + 1

        ws[f"A{r}"] = row["Category"]
        ws[f"A{r}"].alignment = left
        ws[f"A{r}"].border = thin

        ws[f"B{r}"] = row["Revenue"]
        ws[f"B{r}"].number_format = '"$"#,##0.00'
        ws[f"B{r}"].alignment = right
        ws[f"B{r}"].border = thin

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15

    wb.save(filename)
    print(f"Report saved: {filename}")


if __name__ == "__main__":
    from kpi_engine import run_kpi_engine
    kpi_results = run_kpi_engine()
    run_reporter(kpi_results)